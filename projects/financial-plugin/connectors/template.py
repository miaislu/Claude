"""
Excel 财务模型模板管理（openpyxl）。
Model Builder 专用。

MVP：4 个工作表
  封面 / 利润表（历史+预测）/ DCF 估值 / 可比估值
"""

from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore")

_ROOT = Path(__file__).parent.parent

# 样式常量
_HEADER_FILL = "003366"   # 深蓝
_LABEL_FILL  = "E8EAF6"   # 浅紫
_WARN_FILL   = "FFF9C4"   # 浅黄（TV > 70% 警告）


def create_model_workbook(
    stock_code: str,
    company_name: str,
    income_history: list[dict],
    income_forecast: list[dict],
    dcf_data: dict[str, Any],
    comps_data: list[dict],
    version: str = "v1.0",
) -> str:
    """
    生成财务模型 Excel 文件。

    income_history: 历史利润表列表（倒序，最新在前），每项为 dict
    income_forecast: 预测期利润表列表（正序），每项为 dict
    dcf_data: DCF 计算结果 dict
    comps_data: 可比公司估值倍数列表
    返回：生成的 Excel 文件路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1: 封面
    _write_cover(wb.active, stock_code, company_name, version)
    wb.active.title = "封面"

    # Sheet 2: 利润表
    ws_income = wb.create_sheet("利润表")
    _write_income(ws_income, income_history, income_forecast, stock_code)

    # Sheet 3: DCF 估值
    ws_dcf = wb.create_sheet("DCF估值")
    _write_dcf(ws_dcf, dcf_data, stock_code)

    # Sheet 4: 可比估值
    ws_comps = wb.create_sheet("可比估值")
    _write_comps(ws_comps, comps_data, dcf_data, stock_code)

    # 保存
    storage = _ROOT / "storage"
    storage.mkdir(exist_ok=True)
    from datetime import datetime
    date_str = datetime.now().strftime("%Y%m%d")
    path = storage / f"{stock_code}_model_{version}_{date_str}.xlsx"
    wb.save(str(path))
    return str(path)


def _write_cover(ws: Any, code: str, name: str, version: str) -> None:
    from openpyxl.styles import Font, Alignment
    ws["B2"] = f"{name}（{code}）财务分析模型"
    ws["B2"].font = Font(bold=True, size=16)
    ws["B4"] = f"版本：{version}"
    ws["B5"] = "数据来源：akshare / 公司公告"
    from datetime import datetime
    ws["B6"] = f"生成日期：{datetime.now().strftime('%Y-%m-%d')}"
    ws.column_dimensions["B"].width = 40


def _write_income(
    ws: Any,
    history: list[dict],
    forecast: list[dict],
    code: str,
) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # 关键行项目（中文字段名 → 展示标签）
    rows = [
        ("营业收入",     "营业收入"),
        ("营业成本",     "营业成本"),
        ("毛利润",       "毛利润"),
        ("销售费用",     "销售费用"),
        ("管理费用",     "管理费用"),
        ("研发费用",     "研发费用"),
        ("财务费用",     "财务费用"),
        ("营业利润",     "营业利润"),
        ("净利润",       "净利润"),
        ("扣除非经常性损益后的净利润", "扣非净利润"),
    ]

    # 标题行：行名 + 历史期（倒序变正序显示）+ 预测期
    periods_hist = [str(h.get("报告日", ""))[:8] for h in reversed(history)]
    periods_fore = [str(f.get("period", ""))[:7] + "E" for f in forecast]
    all_periods  = periods_hist + periods_fore

    # Row 1: 空 + 期间
    ws.cell(1, 1, "项目（元）").font = Font(bold=True)
    for j, p in enumerate(all_periods, 2):
        cell = ws.cell(1, j, p)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)

    # 数据行
    hist_list = list(reversed(history))
    for i, (field, label) in enumerate(rows, 2):
        ws.cell(i, 1, label)
        # 历史
        for j, rec in enumerate(hist_list, 2):
            ws.cell(i, j, _fmt(rec.get(field)))
        # 预测（来自 forecast dict 的 field 或 computed）
        for j, rec in enumerate(forecast, 2 + len(hist_list)):
            ws.cell(i, j, _fmt(rec.get(field)))

    ws.column_dimensions["A"].width = 20
    for j in range(2, len(all_periods) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14


def _write_dcf(ws: Any, dcf: dict[str, Any], code: str) -> None:
    from openpyxl.styles import Font, PatternFill

    rows = [
        ("WACC",        f"{(dcf.get('wacc') or 0):.2%}"),
        ("永续增长率 g", f"{(dcf.get('g') or 0):.2%}"),
        ("预测期",      f"{dcf.get('forecast_years', 5)} 年"),
        ("企业价值 EV",  _fmt_large(dcf.get("ev"))),
        ("净负债",       _fmt_large(dcf.get("net_debt"))),
        ("股权价值",     _fmt_large(dcf.get("equity_value"))),
        ("总股本（亿股）", f"{(dcf.get('shares') or 0) / 1e8:.2f}"),
        ("DCF 目标价",   f"¥{(dcf.get('dcf_price') or 0):.2f}"),
        ("当前股价",     f"¥{(dcf.get('current_price') or 0):.2f}"),
        ("上行空间",     f"{(dcf.get('upside_pct') or 0):+.1f}%"),
        ("终值/EV",      f"{(dcf.get('tv_ev_ratio') or 0):.1%}"),
    ]

    ws.cell(1, 1, "DCF 估值摘要").font = Font(bold=True, size=13)
    for i, (label, val) in enumerate(rows, 3):
        ws.cell(i, 1, label)
        cell = ws.cell(i, 2, val)
        # TV/EV > 70% 高亮
        if label == "终值/EV" and dcf.get("tv_ev_ratio", 0) > 0.70:
            cell.fill = PatternFill("solid", fgColor=_WARN_FILL)
            ws.cell(i, 3, "⚠ 终值占比过高")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def _write_comps(ws: Any, comps: list[dict], dcf: dict, code: str) -> None:
    from openpyxl.styles import Font, PatternFill

    headers = ["代码", "PE(TTM)", "PB", "说明"]
    ws.cell(1, 1, "可比公司估值").font = Font(bold=True, size=13)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(3, j, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)

    for i, c in enumerate(comps, 4):
        ws.cell(i, 1, c.get("code", ""))
        ws.cell(i, 2, c.get("pe_ttm", "N/A"))
        ws.cell(i, 3, c.get("pb", "N/A"))
        ws.cell(i, 4, c.get("reason", ""))

    # 综合估值摘要
    row_sum = len(comps) + 6
    ws.cell(row_sum, 1, "DCF 目标价").font = Font(bold=True)
    ws.cell(row_sum, 2, f"¥{(dcf.get('dcf_price') or 0):.2f}")
    ws.cell(row_sum + 1, 1, "可比估值目标价").font = Font(bold=True)
    comps_price = dcf.get("comps_price")
    ws.cell(row_sum + 1, 2, f"¥{comps_price:.2f}" if comps_price else "N/A")
    ws.cell(row_sum + 2, 1, "综合目标价（5:5）").font = Font(bold=True)
    blended = dcf.get("blended_price")
    ws.cell(row_sum + 2, 2, f"¥{blended:.2f}" if blended else "N/A")

    for col, w in zip("ABCD", [10, 12, 12, 30]):
        ws.column_dimensions[col].width = w


# ──────────────────────────────────────────
# 辅助格式化
# ──────────────────────────────────────────

def _fmt(v: Any) -> Any:
    """数值格式化（None → 空，保留原始数字给 Excel）。"""
    if v is None:
        return ""
    try:
        import math
        f = float(v)
        return None if math.isnan(f) else f
    except (ValueError, TypeError):
        return str(v)


def _fmt_large(v: Any) -> str:
    """大数格式化为亿元。"""
    if v is None:
        return "N/A"
    try:
        return f"{float(v) / 1e8:.2f} 亿"
    except (ValueError, TypeError):
        return "N/A"
