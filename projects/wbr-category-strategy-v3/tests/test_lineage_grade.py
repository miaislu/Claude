"""Unit + e2e tests for scripts/lineage_grade.py"""
from __future__ import annotations

import json
import subprocess
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from _testutil import FIXTURES_DIR, SCRIPTS_DIR

from wbr_engine.lineage.grade import parse_threshold, grade_prediction, render_grading_md


def make_test_df(values: dict, week_cols=('W14', 'W15')):
    """构造与 common.load_wbr_excel 输出兼容的 DataFrame"""
    rows = []
    for metric, vals in values.items():
        row = {'group': 'Output', 'metric': metric, 'unit': '亿'}
        for i, w in enumerate(week_cols):
            row[w] = vals[i] if i < len(vals) else None
        rows.append(row)
    return pd.DataFrame(rows)


def write_test_excel(path: Path, values: dict):
    """生成与 common.load_wbr_excel 期望的格式兼容的 .xlsx"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'WBR 周粒度数据(测试)'
    ws['A2'] = ''  # 空行
    # 表头行
    headers = ['第一分组标题', '指标', '单位', 'W14##04.06-04.12', 'W15##04.13-04.19']
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=h)
    # 数据行
    for r_idx, (metric, vals) in enumerate(values.items(), start=4):
        ws.cell(row=r_idx, column=1, value='Output')
        ws.cell(row=r_idx, column=2, value=metric)
        ws.cell(row=r_idx, column=3, value='亿')
        ws.cell(row=r_idx, column=4, value=vals[0])
        ws.cell(row=r_idx, column=5, value=vals[1])
    wb.save(str(path))


class TestParseThreshold(unittest.TestCase):
    def test_absolute_gte(self):
        t = parse_threshold('>=2.0亿')
        self.assertEqual(t, {'kind': 'absolute', 'op': '>=', 'value': 2.0, 'unit': '亿'})

    def test_absolute_lt(self):
        t = parse_threshold('< 3.0%')
        self.assertEqual(t['kind'], 'absolute')
        self.assertEqual(t['op'], '<')
        self.assertEqual(t['value'], 3.0)
        self.assertEqual(t['unit'], '%')

    def test_unicode_gte(self):
        t = parse_threshold('≥ 2.0亿')
        self.assertEqual(t['op'], '>=')
        self.assertEqual(t['value'], 2.0)

    def test_relative_pct(self):
        t = parse_threshold('+5%')
        self.assertEqual(t['kind'], 'relative_pct')
        self.assertEqual(t['sign'], '+')
        self.assertEqual(t['value'], 5)

    def test_relative_pp(self):
        t = parse_threshold('-8pp')
        self.assertEqual(t['kind'], 'relative_pp')
        self.assertEqual(t['sign'], '-')

    def test_semantic(self):
        t = parse_threshold('回升')
        self.assertEqual(t['kind'], 'semantic')
        self.assertEqual(t['verb'], '回升')

    def test_unparseable(self):
        self.assertIsNone(parse_threshold('随便写'))


class TestGradePrediction(unittest.TestCase):
    def test_achieved_absolute_gte(self):
        df = make_test_df({'啤酒消费GTV': (1.8, 2.13)})
        pred = {'id': 'p1', 'metric': '啤酒消费GTV', 'threshold': '>=2.0亿', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '✅ achieved')
        self.assertAlmostEqual(r['actual'], 2.13)

    def test_missed_absolute_gte(self):
        df = make_test_df({'啤酒消费GTV': (1.8, 1.95)})
        pred = {'id': 'p1', 'metric': '啤酒消费GTV', 'threshold': '>=2.0亿', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '❌ missed')

    def test_achieved_absolute_lt(self):
        df = make_test_df({'风险水位': (3.2, 2.8)})
        pred = {'id': 'p2', 'metric': '风险水位', 'threshold': '< 3.0%', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '✅ achieved')

    def test_data_missing_metric_not_found(self):
        df = make_test_df({'啤酒消费GTV': (1.8, 2.13)})
        pred = {'id': 'p3', 'metric': '不存在的指标', 'threshold': '>=1', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '⏸ data_missing')

    def test_data_missing_unparseable_threshold(self):
        df = make_test_df({'啤酒消费GTV': (1.8, 2.13)})
        pred = {'id': 'p4', 'metric': '啤酒消费GTV', 'threshold': '一些随便写的', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '⏸ data_missing')

    def test_relative_pct_achieved(self):
        df = make_test_df({'白酒新客数': (4.0, 4.3)})  # +7.5%
        pred = {'id': 'p5', 'metric': '白酒新客数', 'threshold': '+5%', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '✅ achieved')

    def test_relative_pct_missed(self):
        df = make_test_df({'白酒新客数': (4.0, 3.7)})  # -7.5%
        pred = {'id': 'p5', 'metric': '白酒新客数', 'threshold': '+5%', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '❌ missed')

    def test_semantic_huisheng(self):
        df = make_test_df({'啤酒消费GTV': (1.8, 2.1)})
        pred = {'id': 'p6', 'metric': '啤酒消费GTV', 'threshold': '回升', 'by_week': 'W15'}
        r = grade_prediction(pred, df)
        self.assertEqual(r['status'], '✅ achieved')


class TestRenderGradingMd(unittest.TestCase):
    def test_sections_appear(self):
        lineage = {
            'week': 'W14',
            'predictions': [
                {'id': 'p1', 'metric': '啤酒GTV', 'threshold': '>=2亿'},
                {'id': 'p2', 'metric': '风险水位', 'threshold': '<3%'},
            ],
        }
        results = [
            {'id': 'p1', 'status': '✅ achieved', 'actual': 2.1, 'detail': 'detail1'},
            {'id': 'p2', 'status': '❌ missed', 'actual': 3.5, 'detail': 'detail2'},
        ]
        md = render_grading_md(lineage, results)
        self.assertIn('已达成', md)
        self.assertIn('偏离', md)
        self.assertIn('p1', md)
        self.assertIn('p2', md)


class TestE2E(unittest.TestCase):
    """端到端 smoke:真实 Excel → lineage_grade.py → grading.md"""

    def test_full_pipeline_real_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            # 1. 写 Excel
            xlsx = tmp_path / 'W15.xlsx'
            write_test_excel(xlsx, {
                '啤酒消费GTV': (1.8, 2.13),     # ≥ 2.0 → achieved
                '风险水位':    (3.2, 2.8),     # < 3.0 → achieved
                '白酒新客数':  (4.0, 3.7),     # +5% → missed (实际 -7.5%)
            })
            # 2. 写 lineage_prev.json
            lineage_in = tmp_path / 'lineage_prev.json'
            lineage_in.write_text(json.dumps({
                'week': 'W14',
                'predictions': [
                    {'id': 'pred_W14_001', 'metric': '啤酒消费GTV', 'threshold': '>=2.0亿', 'by_week': 'W15'},
                    {'id': 'pred_W14_002', 'metric': '风险水位', 'threshold': '< 3.0%', 'by_week': 'W15'},
                    {'id': 'pred_W14_003', 'metric': '白酒新客数', 'threshold': '+5%', 'by_week': 'W15'},
                    {'id': 'pred_W14_004', 'metric': '不存在的指标', 'threshold': '>=1', 'by_week': 'W15'},
                ],
            }, ensure_ascii=False), encoding='utf-8')

            grading_md = tmp_path / 'grading.md'
            grading_json = tmp_path / 'grading.json'

            result = subprocess.run(
                ['python3', str(SCRIPTS_DIR / 'lineage_grade.py'),
                 '--lineage', str(lineage_in),
                 '--indicator-data', str(xlsx),
                 '--output', str(grading_md),
                 '--output-json', str(grading_json)],
                capture_output=True, text=True
            )
            self.assertEqual(result.returncode, 0, result.stderr)

            # 3. 验证产出
            md_text = grading_md.read_text(encoding='utf-8')
            self.assertIn('已达成 (2/4)', md_text)
            self.assertIn('偏离 (1/4)', md_text)
            self.assertIn('数据未回流 (1/4)', md_text)
            self.assertIn('pred_W14_001', md_text)
            self.assertIn('pred_W14_003', md_text)  # missed
            self.assertIn('pred_W14_004', md_text)  # data_missing

            grading_data = json.loads(grading_json.read_text(encoding='utf-8'))
            statuses = [r['status'] for r in grading_data['results']]
            self.assertEqual(sorted(statuses), sorted(['✅ achieved', '✅ achieved', '❌ missed', '⏸ data_missing']))


if __name__ == '__main__':
    unittest.main()
